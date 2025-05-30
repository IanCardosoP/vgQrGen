"""
Módulo Gestor de Datos Excel.

Este módulo proporciona funcionalidad para leer y procesar archivos Excel
que contienen información de redes WiFi.
"""

import os
import logging
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from typing import Optional, List, Dict, Tuple
from dataclasses import dataclass
from .qr_manager import WiFiCredentials
from ..utils.logging_utils import LogManager
from ..utils.path_utils import resource_path

logger = LogManager.get_logger(__name__)

@dataclass
class ExcelColumns:
    """Clase de datos para almacenar índices de columnas de Excel."""
    room: int
    ssid: int
    password: Optional[int] = None
    encryption: Optional[int] = None
    property_type: Optional[int] = None

class ExcelManager:
    """Gestiona operaciones de archivo Excel y extracción de datos."""
    
    # Palabras clave para auto-detectar encabezados de columnas
    COLUMN_KEYWORDS = {
        'room': ['room', 'habitacion', 'habitación', 'number', 'número', 'hab', 'cuarto', 'villa'],
        'ssid': ['ssid', 'network', 'red', 'wifi', 'nombre', 'name', 'net'],
        'password': ['password', 'contraseña', 'pass', 'key', 'clave', 'pwd', 'contrasena'],
        'encryption': ['security', 'encryption', 'seguridad', 'encriptación', 'type', 'tipo', 'encriptacion'],
        'property': ['property', 'propiedad', 'hotel', 'region', 'zona', 'lugar', 'site']
    }
    
    def __init__(self, file_path: str):
        """
        Inicializar Gestor de Excel.
        
        Args:
            file_path (str): Ruta al archivo Excel
        """        
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        self.columns = None
        self.columns_detected = False  # Indicador si las columnas fueron detectadas (automática o manualmente)
        
    def validate_file(self) -> Tuple[bool, str]:
        """
        Validar el archivo Excel antes de intentar cargarlo.
        
        Returns:
            Tuple[bool, str]: (es_válido, mensaje_error)
        """
        logger.debug(f"Iniciando validación del archivo: {self.file_path}")
        
        if not self.file_path:
            logger.error("No se proporcionó ruta de archivo para validación")
            return False, "No se proporcionó ruta de archivo"
            
        if not os.path.exists(self.file_path):
            logger.error(f"Archivo no encontrado: {self.file_path}")
            return False, f"Archivo no encontrado: {self.file_path}"
            
        if not self.file_path.lower().endswith(('.xlsx', '.xls')):
            logger.error(f"Extensión de archivo inválida: {self.file_path}. Debe ser .xlsx o .xls")
            return False, "El archivo debe ser un archivo Excel (.xlsx o .xls)"
        
        file_size = os.path.getsize(self.file_path)
        logger.debug(f"Tamaño del archivo: {file_size} bytes")
            
        try:
            # Intentar abrir el archivo para verificar que no está corrupto
            with open(self.file_path, 'rb') as f:
                f.read(10)  # Leer primeros bytes
            logger.debug(f"Archivo validado exitosamente: {self.file_path}")
            return True, ""
        except PermissionError:
            error_msg = f"Error de permisos al acceder al archivo: {self.file_path}"
            logger.error(error_msg)
            return False, error_msg
        except Exception as e:
            error_msg = f"El archivo no es accesible: {str(e)}"
            logger.error(f"{error_msg}. Excepción: {type(e).__name__}")
            return False, error_msg
    
    def load_workbook(self) -> Tuple[bool, str]:
        """
        Cargar solo el libro de Excel, sin seleccionar una hoja.
        
        Returns:
            Tuple[bool, str]: (éxito, mensaje_error)
        """
        # Validar archivo primero
        is_valid, error_msg = self.validate_file()
        if not is_valid:
            logger.error(f"Falló la validación del archivo: {error_msg}")
            return False, error_msg
        
        logger.info(f"Iniciando carga del libro Excel: {self.file_path}")
        try:
            logger.debug(f"Abriendo archivo Excel con opción data_only=True")
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            
            # Registro de información sobre el libro
            sheet_count = len(self.workbook.sheetnames)
            logger.info(f"Libro Excel cargado exitosamente: {self.file_path}")
            logger.info(f"El libro contiene {sheet_count} hoja(s): {', '.join(self.workbook.sheetnames)}")
            
            if not self.workbook.sheetnames:
                logger.error("El archivo Excel no contiene hojas")
                return False, "El archivo Excel no contiene hojas"
                
            return True, ""
            
        except InvalidFileException:
            error_msg = "Formato de archivo Excel inválido"
            logger.error(f"{error_msg}. El archivo {self.file_path} no es un archivo Excel válido o está dañado.")
            return False, error_msg
        except PermissionError:
            error_msg = "Error de permisos al acceder al archivo"
            logger.error(f"{error_msg}. Verifique que el archivo {self.file_path} no esté abierto en otro programa.")
            return False, error_msg
        except Exception as e:
            error_msg = f"Error al cargar el libro: {str(e)}"
            logger.error(f"{error_msg}. Excepción: {type(e).__name__}. Detalles: {str(e)}")
            return False, error_msg
    
    def get_sheet_names(self) -> List[str]:
        """
        Obtener lista de nombres de hojas disponibles.
        
        Returns:
            List[str]: Lista de nombres de hojas
        """
        if self.workbook:
            return self.workbook.sheetnames
        return []
    
    def set_active_sheet(self, sheet_name: str) -> Tuple[bool, str]:
        """
        Establecer la hoja activa e intentar detectar columnas.
        
        Args:
            sheet_name (str): Nombre de la hoja a usar
            
        Returns:
            Tuple[bool, str]: (éxito, mensaje_error)
        """
        if not self.workbook:
            logger.error("Intento de seleccionar hoja sin libro cargado")
            return False, "No hay libro cargado"
            
        logger.info(f"Intentando seleccionar hoja: '{sheet_name}'")
        
        if sheet_name not in self.workbook.sheetnames:
            logger.error(f"Hoja '{sheet_name}' no encontrada en el libro. Hojas disponibles: {', '.join(self.workbook.sheetnames)}")
            return False, f"Hoja '{sheet_name}' no encontrada"
            
        logger.debug(f"Seleccionando hoja: '{sheet_name}'")
        self.sheet = self.workbook[sheet_name]
        
        # Verificar si la hoja está vacía
        max_row = self.sheet.max_row
        max_col = self.sheet.max_column
        logger.debug(f"La hoja '{sheet_name}' tiene {max_row} filas y {max_col} columnas")
        
        if max_row < 2:  # Necesita al menos fila de encabezado y una fila de datos
            logger.warning(f"La hoja '{sheet_name}' está vacía o solo contiene encabezados (filas: {max_row})")
            return False, "La hoja seleccionada parece estar vacía o solo contiene encabezados"
            
        # Intentar detectar columnas de la fila de encabezado
        logger.info(f"Intentando detectar columnas automáticamente en la hoja '{sheet_name}'")
        self.columns = self._detect_columns()
        
        if not self.columns:
            logger.warning(f"No se pudieron detectar columnas requeridas en la hoja '{sheet_name}'. Se requiere configuración manual.")
            return False, "No se encuentran las columnas requeridas (habitación y SSID). Por favor, utilice la configuración manual."
        
        self.columns_detected = True  # Actualizar indicador de detección de columnas
        
        # Registrar información sobre las columnas detectadas
        column_info = []
        if hasattr(self.columns, 'room'):
            column_info.append(f"room: {self.columns.room}")
        if hasattr(self.columns, 'ssid'):
            column_info.append(f"ssid: {self.columns.ssid}")
        if hasattr(self.columns, 'password') and self.columns.password is not None:
            column_info.append(f"password: {self.columns.password}")
        if hasattr(self.columns, 'encryption') and self.columns.encryption is not None:
            column_info.append(f"encryption: {self.columns.encryption}")
        if hasattr(self.columns, 'property_type') and self.columns.property_type is not None:
            column_info.append(f"property_type: {self.columns.property_type}")
            
        logger.info(f"Columnas detectadas en la hoja '{sheet_name}': {', '.join(column_info)}")
        logger.info(f"Hoja '{sheet_name}' cargada exitosamente")
            
        return True, ""
            
    def _detect_columns(self) -> Optional[ExcelColumns]:
        """
        Detectar automáticamente columnas relevantes de la fila de encabezado.
        
        Returns:
            Optional[ExcelColumns]: Índices de columnas si se encuentran
        """
        if not self.sheet:
            return None
            
        column_indices = {
            'room': None,
            'ssid': None,
            'password': None,
            'encryption': None,
            'property': None
        }
        
        # Obtener fila de encabezado (primera fila)
        header_row = list(self.sheet.iter_rows(min_row=1, max_row=1))[0]
        
        # Buscar en fila de encabezado nombres de columnas
        for idx, cell in enumerate(header_row):
            if not cell.value:
                continue
                
            cell_value = str(cell.value).lower().strip()
            
            # Verificar coincidencias exactas primero
            for col_type, keywords in self.COLUMN_KEYWORDS.items():
                if cell_value in keywords:
                    column_indices[col_type] = idx
                    break
                    
            # Si no hay coincidencia exacta, verificar coincidencias parciales
            if all(v is not None for v in column_indices.values()):
                continue
                
            for col_type, keywords in self.COLUMN_KEYWORDS.items():
                if column_indices[col_type] is not None:
                    continue
                if any(keyword in cell_value for keyword in keywords):
                    column_indices[col_type] = idx
                    break
        
        # Registrar columnas encontradas y faltantes
        found_cols = [k for k, v in column_indices.items() if v is not None]
        missing_cols = [k for k, v in column_indices.items() if v is None]
        logger.info(f"Columnas encontradas en fila de encabezado: {', '.join(found_cols)}")
        if missing_cols:
            logger.warning(f"Columnas faltantes en fila de encabezado: {', '.join(missing_cols)}")
        
        # Verificar que se encontraron las columnas requeridas
        if column_indices['room'] is None or column_indices['ssid'] is None:
            logger.error("Columnas requeridas 'habitación' y 'ssid' no encontradas en fila de encabezado")
            return None
            
        return ExcelColumns(
            room=column_indices['room'],
            ssid=column_indices['ssid'],
            password=column_indices.get('password'),
            encryption=column_indices.get('encryption'),
            property_type=column_indices.get('property')
        )

    def get_room_data(self, room_number: str) -> Optional[WiFiCredentials]:
        """
        Obtener credenciales WiFi para una habitación específica.
        
        Args:
            room_number (str): Número de habitación a buscar
            
        Returns:
            Optional[WiFiCredentials]: Credenciales WiFi de la habitación si se encuentra
        """
        if not self.sheet or not self.columns:
            logger.error(f"Intento de buscar habitación '{room_number}' sin hoja activa o columnas configuradas")
            return None
            
        room_number = str(room_number).strip().upper()
        logger.info(f"Buscando habitación: {room_number}")
        
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if str(row[self.columns.room]).strip().upper() == room_number:
                logger.info(f"Habitación {room_number} encontrada")
                
                # Si hay columna de encryption y tiene valor, usarlo; si no, devolver None
                encryption = None
                if self.columns.encryption is not None and row[self.columns.encryption]:
                    encryption = str(row[self.columns.encryption]).strip()
                    logger.debug(f"Encriptación encontrada: {encryption}")
                else:
                    logger.debug("No se encontró valor de encriptación en la fila")
                    
                # Preparar valores de credenciales
                ssid = str(row[self.columns.ssid]).strip()
                password = str(row[self.columns.password]).strip() if self.columns.password is not None and row[self.columns.password] else None
                property_type = str(row[self.columns.property_type]).strip() if self.columns.property_type is not None and row[self.columns.property_type] else None
                
                logger.info(f"Credenciales para habitación {room_number} - SSID: {ssid}, Encriptación: {encryption}, Propiedad: {property_type}")
                
                return WiFiCredentials(
                    ssid=ssid,
                    password=password,
                    encryption=encryption,  # Puede ser None
                    property_type=property_type
                )
        
        logger.warning(f"Habitación {room_number} no encontrada en la hoja activa")
        return None
        
    def get_all_rooms(self) -> List[WiFiCredentials]:
        """
        Obtener credenciales WiFi para todas las habitaciones en la hoja.
        
        Returns:
            List[WiFiCredentials]: Lista de todas las credenciales de habitaciones
        """
        if not self.sheet or not self.columns:
            logger.error("Intento de obtener todas las habitaciones sin hoja activa o columnas configuradas")
            return []
            
        logger.info("Iniciando extracción de todas las habitaciones en la hoja")
        credentials = []
        row_count = 0
        
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            row_count += 1
            if not row[self.columns.room] or not row[self.columns.ssid]:
                logger.debug(f"Ignorando fila {row_count+1} por valores de habitación o SSID vacíos")
                continue
                
            room = str(row[self.columns.room]).strip()
            ssid = str(row[self.columns.ssid]).strip()
            
            # Si hay columna de encryption y tiene valor, usarlo; si no, devolver None
            encryption = None
            if self.columns.encryption is not None and row[self.columns.encryption]:
                encryption = str(row[self.columns.encryption]).strip()
                
            password = str(row[self.columns.password]).strip() if self.columns.password is not None and row[self.columns.password] else None
            property_type = str(row[self.columns.property_type]).strip() if self.columns.property_type is not None and row[self.columns.property_type] else None
            
            logger.debug(f"Encontrada habitación {room} - SSID: {ssid}, Encriptación: {encryption}, Propiedad: {property_type}")
                
            cred = WiFiCredentials(
                ssid=ssid,
                password=password,
                encryption=encryption,  # Puede ser None
                property_type=property_type
            )
            credentials.append(cred)
            
        logger.info(f"Total de habitaciones encontradas: {len(credentials)}")
        return credentials
        
    def set_columns_manually(self, column_indices: Dict[str, int]) -> bool:
        """
        Establecer índices de columnas manualmente.
        
        Args:
            column_indices (Dict[str, int]): Diccionario que mapea nombres de columnas a índices
            
        Returns:
            bool: True si fue exitoso
        """
        try:
            self.columns = ExcelColumns(
                room=column_indices['room'],
                ssid=column_indices['ssid'],
                password=column_indices.get('password'),
                encryption=column_indices.get('encryption'),
                property_type=column_indices.get('property_type')
            )
            self.columns_detected = True  # Actualizar indicador de detección de columnas
            return True
        except Exception as e:
            logger.error(f"Error al establecer columnas manualmente: {str(e)}")
            return False