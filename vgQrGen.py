import argparse
import sys
import segno
import os
import openpyxl
from io import BytesIO
from typing import Optional, Tuple, Union
import tqdm
from PIL import Image, ImageDraw, ImageFont
from segno import helpers

# Configuración global
DEFAULT_SOURCE_FILE = "source.xlsx"
DEFAULT_SHEET_NAME = "ANTENAS ARUBA ESTANCIA 2025"
DEFAULT_ENCRIPTION = "WPA2"

# Rutas de logos
LOGO_PATH_VILLAGROUP = "assets/logo_villagroup.png"
LOGO_PATH_VILLA_ESTANCIA = "assets/logo_villa_estancia.png"

def main():
    # Configuración de argumentos
    parser = argparse.ArgumentParser(
        description="Script generador de QR para Wifi. Usa un único parámetro para ejecutar diferentes funcionalidades."
    )
    parser.add_argument(
        "habitacion", nargs="?", type=str,
        help="Especifica la habitación, ejemplo: 1102A."
    )
    parser.add_argument(
        "--all", action="store_true",
        help="Generar para todas las habitaciones disponibles."
    )
    parser.add_argument(
        "--new", action="store_true",
        help="Formulario para crear una nueva habitación."
    )

    args = parser.parse_args()
    argumentos = sum([bool(args.habitacion), args.all, args.new])

    if argumentos > 1:
        print("Error: Solo puedes usar un argumento a la vez.")
        return 1

    try:
        if args.habitacion:
            print(f"Buscando habitación {args.habitacion}...")
            sheet = load_source_file()
            if not sheet:
                return 1
                
            ssid, password = search_for_room(args.habitacion, sheet)
            if ssid and password:
                propiedad = validar_propiedad(args.propiedad)
                encryption = DEFAULT_ENCRIPTION
                try:
                    generar_qr(ssid, password, encryption, propiedad)
                    print(f"Código QR generado exitosamente: {output_path}")
                except Exception as e:
                    print(f"Error al generar el QR: {str(e)}")
                    return 1
                
            else:
                return 1
                
        elif args.all:
            print("Iterando hoja y generando para todas las habitaciones...")
            wb, sheet = load_source_file()
            if sheet:
                generar_todo(sheet, DEFAULT_ENCRIPTION)
                
        elif args.new:
            print("Generando QR para nueva red.")
            ssid = leer_ssid()
            encryption = leer_seguridad()
            password = leer_pass()
            output_path = generar_qr(ssid, password, encryption)
            print(f"Código QR generado exitosamente: {output_path}")
            
        else:
            mostrar_ayuda()
            
    except Exception as e:
        print(f"Error inesperado: {str(e)}")
        return 1
        
    return 0

def mostrar_ayuda():
    print("Uso:")
    print("  python vgQrGen.py [habitacion] | --all | --new")
    print("Ejemplos:")
    print("  python vgQrGen.py 1102A       # Buscar y generar para habitación específica")
    print("  python vgQrGen.py --all       # Generar QR para todas las habitaciones")
    print("  python vgQrGen.py --new       # Generar QR para nueva habitación")
    print("\nUsa --help para más información.")
        
def load_source_file(
    file_path: str = DEFAULT_SOURCE_FILE, 
    sheet_name: str = DEFAULT_SHEET_NAME,
    data_only: bool = True
) -> Tuple[Optional[openpyxl.Workbook], Optional[openpyxl.worksheet.worksheet.Worksheet]]:
    """
    Carga un archivo Excel y devuelve el workbook y la hoja especificada.
    
    Args:
        file_path (str): Ruta del archivo Excel. Por defecto usa DEFAULT_SOURCE_FILE.
        sheet_name (str): Nombre de la hoja a cargar. Por defecto usa DEFAULT_SHEET_NAME.
        data_only (bool): Si True, carga solo valores (no fórmulas). Por defecto True.   
    Returns:
        tuple: (Workbook, Worksheet) o (None, None) si hay error.
    
    Raises:
        FileNotFoundError: Si el archivo no existe.
        ValueError: Si la hoja no existe en el workbook.
    """
    try:
        # Validación básica de parámetros
        if not file_path or not isinstance(file_path, str):
            raise ValueError("La ruta del archivo debe ser una cadena no vacía")
            
        if not sheet_name or not isinstance(sheet_name, str):
            raise ValueError("El nombre de la hoja debe ser una cadena no vacía")

        # Cargar workbook
        wb = openpyxl.load_workbook(file_path, data_only=data_only)
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            available_sheets = ", ".join(wb.sheetnames)
            raise ValueError(
                f"Hoja '{sheet_name}' no encontrada. Hojas disponibles: {available_sheets}"
            )
        
        sheet = wb[sheet_name]
        return wb, sheet
        
    except FileNotFoundError as fnf_error:
        print(f"Error: Archivo no encontrado - {fnf_error}")
        return None, None
    except ValueError as val_error:
        print(f"Error de valor: {val_error}")
        return None, None
    except openpyxl.utils.exceptions.InvalidFileException as file_error:
        print(f"Error: Archivo Excel inválido - {file_error}")
        return None, None
    except Exception as e:
        print(f"Error inesperado al cargar el archivo: {str(e)}")
        return None, None

def search_for_room(habitacion, sheet):
    """Busca los datos WiFi para una habitación en el Excel.
    
    Args:
        habitacion (str): Nombre/número de la habitación a buscar
        
    Returns:
        tuple: (ssid, password) o (None, None) si no se encuentra
    """
    is_there_habitacion = exist_habitacion(habitacion)
    col_room = get_habitacion_idx()

    is_there_encrypt = exist_encryption()
    col_encrypt = get_encryption_idx()

    is_there_propriety = exist_propriety()
    col_propriety = get_propriety_idx()

    is_there_pss = exist_pss()
    col_passwd = get_pss_idx()

    is_there_ssid = exist_ssid()
    col_ssid = get_ssid_idx()

    try:
        for fila in sheet.iter_rows(
            min_row=1, 
            max_row=sheet.max_row, 
            min_col=1, 
            max_col=sheet.max_column, 
            values_only=True
        ):
            if str(fila[col_room]).strip().upper() == habitacion.strip().upper():              

                # El valor es igual al valor de tal columna si es que la columna existe. Si no, None.                
                sid = fila[col_ssid] if is_there_ssid else None
                pss = fila[col_passwd] if is_there_pss else None
                encrypt = fila[col_encrypt] if is_there_encrypt else DEFAULT_ENCRIPTION
                region = fila[col_propriety] if is_there_propriety else None
             
                
                # Validar que los valores no estén vacíos
                if sid and pss:
                    return 
                    str(ssid).strip(), 
                    str(password).strip(), 
                    str(encryption).strip(),
                    str(propriety).strip()
                
        print(f"No se encontró la habitación '{habitacion}' o faltan datos")
        return None

    except Exception as e:
        print(f"Error al buscar en el archivo Excel: {str(e)}")
        return None, None
    
def search_for_room(habitacion, sheet):
    """Busca los datos WiFi para una habitación en el Excel.
    
    Args:
        habitacion (str): Nombre/número de la habitación a buscar
        sheet (Worksheet): Objeto de hoja Excel de openpyxl
        
    Returns:
        tuple: (ssid, password, encryption, propriety) o (None, None, None, None) si no se encuentra
    """
    # Validación inicial
    if not habitacion or not isinstance(habitacion, str):
        print("El nombre de la habitación no es válido")
        return None, None, None, None
    
    try:
        # Obtener índices de columnas (preferible mover esta lógica fuera de la función)
        col_room = get_habitacion_idx()
        col_ssid = get_ssid_idx() if exist_ssid() else None
        col_passwd = get_pss_idx() if exist_pss() else None
        col_encrypt = get_encryption_idx() if exist_encryption() else None
        col_propriety = get_propriety_idx() if exist_propriety() else None

        # Normalizar búsqueda
        habitacion_buscada = habitacion.strip().upper()
        
        for fila in sheet.iter_rows(
            min_row=2,  # Asumiendo que la fila 1 es el encabezado
            max_row=sheet.max_row, 
            values_only=True
        ):
            # Verificar si la habitación coincide
            if str(fila[col_room]).strip().upper() == habitacion_buscada:
                # Obtener valores con manejo de columnas faltantes
                ssid = str(fila[col_ssid]).strip() if col_ssid is not None and fila[col_ssid] else None
                password = str(fila[col_passwd]).strip() if col_passwd is not None and fila[col_passwd] else None
                encryption = str(fila[col_encrypt]).strip() if col_encrypt is not None and fila[col_encrypt] else DEFAULT_ENCRIPTION
                propriety = str(fila[col_propriety]).strip() if col_propriety is not None and fila[col_propriety] else None
                
                # Validar campos obligatorios
                if not ssid or not password:
                    print(f"Advertencia: Habitación '{habitacion}' encontrada pero SSID o Password están vacíos")
                    continue  # Por si hay múltiples entradas para la misma habitación
                
                return ssid, password, encryption, propriety
        
        print(f"No se encontró la habitación '{habitacion}' o faltan datos obligatorios")
        return None, None, None, None

    except Exception as e:
        print(f"Error al buscar en el archivo Excel: {str(e)}")
        return None, None, None, None
   
def load_qr_string(ssid: str, password: Optional[str], encryption: str, hidden: bool = False) -> str:
    """Genera el string QR para WiFi.
    
    
    Args:
        ssid (str): Nombre de la red WiFi
        password (str): Contraseña de la red WiFi (opcional)
        encryption (str): Tipo de seguridad (WPA/WEP/nopass)
        hidden (bool): Si la red es oculta (opcional)
        
    Returns:
        Retorna un string con el formato para códigos QR WiFi.
        Este formato es el estándar utilizado para generar códigos QR de redes WiFi que pueden ser escaneados por dispositivos móviles para conectarse automáticamente a la red.
        str: String QR para WiFi
    
    Example:            "WIFI:S:nombre_red;T:WPA2;P:contraseña;"
    Example sin pass:   "WIFI:S:nombre_red;T:nopass;"
    """
    
    if encryption == "nopass":
        return f"WIFI:S:{ssid};T:nopass;"
    else:
        return f"WIFI:S:{ssid};T:{encryption};P:{password};"

def load_qr_on_buffer(
    ssid: str,
    password: Optional[str],
    encryption: str,
    hidden: bool = False
) -> BytesIO:
    """
    Genera un código QR para WiFi y lo guarda en un buffer.
    
    Args:
        ssid (str): Nombre de la red WiFi
        password (str): Contraseña de la red WiFi (opcional)
        encryption (str): Tipo de seguridad (WPA/WEP/nopass)
        hidden (bool): Si la red es oculta (opcional)
        
    Returns:
        BytesIO: Buffer conteniendo la imagen del código QR en formato PNG
        
    Raises:
        ValueError: Si los parámetros no son válidos
    """
    # Obtener el string WiFi QR
    wifi_string = load_qr_string(ssid, password, encryption, hidden)
    segno.helpers.make_wifi_data(wifi_string)  
    
    # Guardar QR en buffer
    buffer = BytesIO()
    buffer.save(buffer, kind='png', scale=10)
    buffer.seek(0) # Rewind the buffer
    
    return buffer

def validar_propiedad(propiedad: str) -> str:
    """Valida la propiedad para cargar el logo y retorna el identificador correspondiente.
    
    Args:
        propiedad (str): Identificador de la propiedad ('VLEV', 'VDPF', 'VLE', 'VPF', 'VG')
        
    Returns:
        str: 'VLEV' o 'VG' dependiendo de la propiedad
        
    Raises:
        ValueError: Si la propiedad no es válida
    """
    propiedad = propiedad.upper()
    if propiedad in ('VLEV', 'VLE'):
        return 'VLEV'
    elif propiedad in ('VDPF', 'VG', 'VDP'):
        return 'VG'
    else:
        raise ValueError("Propiedad no válida.")

def load_logo(propiedad: str) -> Optional[Image.Image]:
    """Carga y valida el logotipo en RGBA según la propiedad.
    
    Args:
        propiedad (str): Identificador de la propiedad ('villagroup' o 'villa_estancia')
        
    Returns:
        Optional[Image.Image]: Imagen del logo o None si hay error
        
    Raises:
        ValueError: Si la propiedad no es válida
    """
    if propiedad == ('VLEV'):
        logo_path = LOGO_PATH_VILLA_ESTANCIA
    else:
        logo_path = LOGO_PATH_VILLAGROUP

    try:
        logo_path = validar_propiedad(propiedad)
        logo_img = Image.open(logo_path)
        # Asegurar que el logo tiene canal alfa para transparencia
        if logo_img.mode != 'RGBA':
            logo_img = logo_img.convert('RGBA')
        return logo_img
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo del logo en {logo_path}")
        return None
    except Exception as e:
        print(f"Error al abrir el logo: {e}")
        return None

def calculate_logo_position(qr_img: Image.Image, logo_img: Image.Image) -> Tuple[Image.Image, Tuple[int, int]]:
    """Calcula el tamaño y posición óptimos del logo en el QR.
    
    Args:
        qr_img (Image.Image): Imagen del código QR
        logo_img (Image.Image): Imagen del logo from load_logo()
        
    Returns:
        Tuple[Image.Image, Tuple[int, int]]: Logo redimensionado y tupla con posición (x, y)
    """
    qr_width, qr_height = qr_img.size
    # El logo no debe exceder el 25% del tamaño del QR
    max_logo_size = min(qr_width, qr_height) // 4
    
    # Crear una copia del logo para no modificar el original
    logo_resized = logo_img.copy()
    logo_width, logo_height = logo_resized.size
    
    # Redimensionar si es necesario
    if logo_width > max_logo_size or logo_height > max_logo_size:
        logo_resized.thumbnail((max_logo_size, max_logo_size), Image.Resampling.LANCZOS)
        logo_width, logo_height = logo_resized.size
    
    # Calcular posición central
    x_pos = (qr_width - logo_width) // 2
    y_pos = (qr_height - logo_height) // 2
    
    return logo_resized, (x_pos, y_pos)

def draw_logo_on_qr(buffer: BytesIO, propiedad: str) -> BytesIO:
    """Dibuja un logo en el código QR.
    
    Args:
        buffer (BytesIO): Buffer conteniendo la imagen del QR
        propiedad (str): Identificador de la propiedad para el logo
        
    Returns:
        BytesIO: Buffer con la imagen del QR con logo
        
    Raises:
        ValueError: Si la propiedad no es válida o hay error en el proceso
    """
    try:
        # Abrir el QR y asegurar canal alfa
        qr_img = Image.open(buffer)
        if qr_img.mode != 'RGBA':
            qr_img = qr_img.convert('RGBA')
            
        # Cargar y validar logo
        logo_img = load_logo(propiedad)
        if not logo_img:
            raise ValueError(f"No se pudo cargar el logo para la propiedad {propiedad} en la funcion draw_logo_on_qr()")
            
        # Calcular posición y tamaño del logo
        logo_resized, (x_pos, y_pos) = calculate_logo_position(qr_img, logo_img)
        
        # Pegar el logo usando el canal alfa como máscara
        qr_img.paste(logo_resized, (x_pos, y_pos), logo_resized)
        
        # Guardar resultado en nuevo buffer
        new_buffer = BytesIO()
        qr_img.save(new_buffer, format='PNG')
        new_buffer.seek(0)
        
        return new_buffer
        
    except Exception as e:
        raise ValueError(f"Error al procesar el QR con logo en draw_logo_on_qr(): {str(e)}")

def cerrar():
    # TODO: Desimplementar la funcion cerrar()
    sys.exit(0)
      
def generar_todo(hoja, encryption: str = DEFAULT_ENCRIPTION, col_room: int = 0, 
                col_password: int = 11, col_ssid: int = 14) -> None:
    """Genera códigos QR para todas las habitaciones en una hoja Excel.
    
    Args:
        hoja: Objeto de hoja Excel (openpyxl)
        encryption: Tipo de seguridad WiFi (WPA/WEP/nopass)
        col_room: Índice columna habitación (0-based)
        col_password: Índice columna contraseña (0-based)
        col_ssid: Índice columna SSID (0-based)
        
    Raises:
        ValueError: Si no hay datos válidos o parámetros incorrectos
    """
    # Validación de parámetros
    if not hoja:
        raise ValueError("El objeto hoja no puede ser None")
    
    if encryption not in ("WPA", "WEP", "WPA2", "nopass"):
        raise ValueError("encryption debe ser WPA, WEP o nopass")
    
    # Obtener filas válidas (no None y con SSID)
    try:
        filas_validas = [
            fila for fila in hoja.iter_rows(min_row=2, values_only=True)
            if fila[col_room] is not None and fila[col_ssid] is not None
        ]
    except Exception as e:
        raise RuntimeError(f"Error al leer datos del Excel: {str(e)}") from e
    
    if not filas_validas:
        print("Advertencia: No se encontraron filas válidas para procesar")
        return

    # Procesamiento con barra de progreso
    errores = 0
    with tqdm.tqdm(filas_validas, desc="Generando QR WiFi", unit="habitación") as barra:
        for fila in barra:
            try:
                room = str(fila[col_room]).strip()
                ssid = str(fila[col_ssid]).strip()
                password = str(fila[col_password]).strip() if fila[col_password] else None
                
                # Validar datos mínimos
                if not ssid:
                    barra.write(f"Advertencia: SSID vacío para habitación {room}")
                    continue
                    
                if encryption != "nopass" and not password:
                    barra.write(f"Advertencia: Contraseña vacía para {ssid} (habitación {room})")
                    continue
                
                # Generar QR
                generar_qr(ssid=ssid, password=password, encryption=encryption)
                barra.set_postfix(habitación=room, estado="OK")
                
            except Exception as e:
                errores += 1
                barra.write(f"Error procesando {room if 'room' in locals() else 'fila'}: {str(e)}")
                barra.set_postfix(habitación=room, estado="ERROR")
                continue
    
    # Resumen final
    print(f"\nProceso completado. Habitaciones procesadas: {len(filas_validas)-errores}, Errores: {errores}")

def leer_ssid() -> str:
    """Solicita y retorna el SSID"""
    ssid = input("Introduce el SSID: ").strip()
    while not ssid:
        print("El SSID no puede estar vacío")
        ssid = input("Introduce el SSID: ").strip()
    return ssid

def leer_seguridad() -> str:
    """Solicita y retorna el tipo de seguridad"""
    encryption = input("Introduce el tipo de seguridad (WPA/WPA2/WEP/nopass) [WPA2]: ").strip() or "WPA2"
    while encryption not in ("WPA", "WPA2", "WEP", "nopass"):
        print("Tipo de seguridad no válido")
        encryption = input("Introduce el tipo de seguridad (WPA/WPA2/WEP/nopass) [WPA2]: ").strip() or "WPA2"
    return encryption

def leer_pass() -> str:
    """Solicita y retorna la contraseña"""
    return input("Introduce la contraseña: ").strip()

def generar_qr(ssid: str, password: str, encryption: str, propiedad: str, hidden: bool = False):
    """Genera un código QR para una red WiFi.
    
    Args:
        ssid (str): Nombre de la red WiFi
        password (str): Contraseña de la red
        encryption (str): Tipo de seguridad (WPA/WPA2/WEP/nopass)
        hidden (bool): Si la red es oculta
        
    Returns:
        str: Ruta del archivo QR generado
        
    Raises:
        ValueError: Si los parámetros no son válidos
    """
    try:
        # Generar QR en buffer
        buffer = load_qr_on_buffer(ssid, password, encryption, hidden)
        
        # Añadir logo al QR (por defecto usamos el logo de Villa Group)
        buffer_with_logo = draw_logo_on_qr(buffer, 'villagroup')
        
        # Crear directorio codes si no existe
        os.makedirs('codes', exist_ok=True)
        
        # Generar nombre de archivo y guardar
        output_path = os.path.join('codes', f'qr_{ssid.replace(" ", "_")}.png')
        with open(output_path, 'wb') as f:
            f.write(buffer_with_logo.getvalue())
            
        
        
    except Exception as e:
        raise ValueError(f"Error al generar el QR: {str(e)}")


def load_qr_as_pil(buffer: BytesIO, add_text_space: bool = False, text_height: int = 60) -> Image.Image:
    """
    Carga un QR desde un buffer como imagen PIL, opcionalmente añadiendo espacio para texto.
    
    Args:
        buffer (BytesIO): Buffer conteniendo la imagen QR
        add_text_space (bool): Si se debe añadir espacio para texto
        text_height (int): Altura del espacio para texto
        
    Returns:
        Image.Image: Imagen PIL del QR
    """
    img_qr = Image.open(buffer)
    
    if not add_text_space:
        return img_qr
        
    width, height = img_qr.size
    new_img = Image.new('RGB', (width, height + text_height), 'white')
    new_img.paste(img_qr, (0, 0))
    return new_img

def configure_font(size: int = 22) -> ImageFont.FreeTypeFont:
    """
    Configura y retorna la fuente para el texto.
    
    Args:
        size (int): Tamaño de la fuente
        
    Returns:
        ImageFont.FreeTypeFont: Objeto fuente configurado
    """
    try:
        font_path = "calibrib.ttf"
        return ImageFont.truetype(font_path, size)
    except:
        return ImageFont.load_default()

def add_text_to_qr(img: Image.Image, texto_superior: str, texto_inferior: str, 
                   font: Optional[ImageFont.FreeTypeFont] = None) -> Image.Image:
    """
    Añade texto centrado a una imagen QR.
    
    Args:
        img (Image.Image): Imagen PIL del QR
        texto_superior (str): Texto a mostrar en la línea superior
        texto_inferior (str): Texto a mostrar en la línea inferior
        font (ImageFont.FreeTypeFont): Fuente a usar (opcional)
        
    Returns:
        Image.Image: Imagen con texto añadido
    """
    if font is None:
        font = configure_font()
        
    draw = ImageDraw.Draw(img)
    width = img.width
    height = img.height - 60  # Altura original del QR
    
    # Texto superior
    text_width = draw.textlength(texto_superior, font=font)
    x_pos = (width - text_width) / 2
    y_pos = height - 20
    draw.text((x_pos, y_pos), texto_superior, font=font, fill="black")
    
    # Texto inferior
    text_width = draw.textlength(texto_inferior, font=font)
    x_pos = (width - text_width) / 2
    y_pos = height + 5
    draw.text((x_pos, y_pos), texto_inferior, font=font, fill="black")
    
    return img

def save_qr_image(img: Image.Image, output_path: str, format: str = 'PNG') -> str:
    """
    Guarda una imagen QR en un archivo.
    
    Args:
        img (Image.Image): Imagen PIL a guardar
        output_path (str): Ruta donde guardar la imagen
        format (str): Formato de imagen (default: PNG)
        
    Returns:
        str: Ruta del archivo guardado
    """
    # Asegurar que el directorio existe
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Guardar imagen
    img.save(output_path, format=format)
    return output_path

def exist_encryption() -> bool:
    """
    Verifica si existe una columna relacionada con encriptación en la primera fila.
    
    Returns:
        bool: True si encuentra una columna de encriptación, False en caso contrario
    """
    wb, sheet = load_source_file()
    if not wb or not sheet:
        return False
        
    # Palabras clave que indican una columna de encriptación
    keywords = ['seguridad', 'encryption', 'encriptado', 'encriptacion', 'security']
    
    # Buscar en la primera fila
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                value = cell.value.lower().strip()
                if any(keyword in value for keyword in keywords):
                    return True
    
    return False

def get_encryption_idx() -> int:
    """
    Obtiene el índice de la columna relacionada con encriptación.
    Debe llamarse solo si exist_encryption() devuelve True.
    
    Returns:
        int: Índice de la columna (0-based)
        
    Raises:
        ValueError: Si no existe columna de encriptación
    """
    wb, sheet = load_source_file()
    if not wb or not sheet:
        raise ValueError("No se pudo cargar el archivo fuente")
        
    keywords = ['seguridad', 'encryption', 'encriptado', 'encriptacion', 'security']
    
    # Buscar en la primera fila
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for idx, cell in enumerate(row):
            if cell.value and isinstance(cell.value, str):
                value = cell.value.lower().strip()
                if any(keyword in value for keyword in keywords):
                    return idx
                    
    raise ValueError("No se encontró columna de encriptación en el fichero fuente")

def exist_propriety() -> bool:
    """
    Verifica si existe una columna relacionada con propiedad en la primera fila.
    
    Returns:
        bool: True si encuentra una columna de proipiedad, False en caso contrario
    """
    wb, sheet = load_source_file()
    if not wb or not sheet:
        return False
        
    # Palabras clave que indican una columna de encriptación
    keywords = ['propiedad', 'region', 'hotel', 'zona']
    
    # Buscar en la primera fila
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                value = cell.value.lower().strip()
                if any(keyword in value for keyword in keywords):
                    return True
    
    return False

def get_propriety_idx() -> int:
    """
    Obtiene el índice de la columna relacionada con propiedad.
    Debe llamarse solo si exist_propriety() devuelve True.
    
    Returns:
        int: Índice de la columna (0-based)
        
    Raises:
        ValueError: Si no existe columna de encriptación
    """
    wb, sheet = load_source_file()
    if not wb or not sheet:
        raise ValueError("No se pudo cargar el archivo fuente")
        
    keywords = ['propiedad', 'region', 'hotel', 'zona']
    
    # Buscar en la primera fila
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for idx, cell in enumerate(row):
            if cell.value and isinstance(cell.value, str):
                value = cell.value.lower().strip()
                if any(keyword in value for keyword in keywords):
                    return idx
                    
    raise ValueError("No se encontró columna de propiedad en el fichero fuente")



if __name__ == "__main__":
    sys.exit(main())