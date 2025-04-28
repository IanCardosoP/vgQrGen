"""
Excel Data Manager Module.

This module provides functionality for reading and processing Excel files
containing WiFi network information.
"""

import os
import logging
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from typing import Optional, List, Dict, Tuple
from dataclasses import dataclass
from .qr_manager import WiFiCredentials
from ..utils.logging_utils import LogManager

logger = LogManager.get_logger(__name__)

@dataclass
class ExcelColumns:
    """Data class for storing Excel column indices."""
    room: int
    ssid: int
    password: Optional[int] = None
    encryption: Optional[int] = None
    property_type: Optional[int] = None

class ExcelManager:
    """Manages Excel file operations and data extraction."""
    
    # Keywords for auto-detecting column headers
    COLUMN_KEYWORDS = {
        'room': ['room', 'habitacion', 'habitación', 'number', 'número', 'hab', 'cuarto', 'villa'],
        'ssid': ['ssid', 'network', 'red', 'wifi', 'nombre', 'name', 'net'],
        'password': ['password', 'contraseña', 'pass', 'key', 'clave', 'pwd', 'contrasena'],
        'encryption': ['security', 'encryption', 'seguridad', 'encriptación', 'type', 'tipo', 'encriptacion'],
        'property': ['property', 'propiedad', 'hotel', 'region', 'zona', 'lugar', 'site']
    }
    
    def __init__(self, file_path: str):
        """
        Initialize Excel Manager.
        
        Args:
            file_path (str): Path to Excel file
        """
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        self.columns = None
        
    def validate_file(self) -> Tuple[bool, str]:
        """
        Validate the Excel file before attempting to load it.
        
        Returns:
            Tuple[bool, str]: (is_valid, error_message)
        """
        if not self.file_path:
            return False, "No file path provided"
            
        if not os.path.exists(self.file_path):
            return False, f"File not found: {self.file_path}"
            
        if not self.file_path.lower().endswith(('.xlsx', '.xls')):
            return False, "File must be an Excel file (.xlsx or .xls)"
            
        try:
            # Try to open the file to check if it's not corrupted
            with open(self.file_path, 'rb') as f:
                f.read(10)  # Read first few bytes
            return True, ""
        except Exception as e:
            return False, f"File is not accessible: {str(e)}"
    
    def load_workbook(self) -> Tuple[bool, str]:
        """
        Load the Excel workbook only, without selecting a sheet.
        
        Returns:
            Tuple[bool, str]: (success, error_message)
        """
        # Validate file first
        is_valid, error_msg = self.validate_file()
        if not is_valid:
            logger.error(f"File validation failed: {error_msg}")
            return False, error_msg
            
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            if not self.workbook.sheetnames:
                return False, "Excel file contains no sheets"
            return True, ""
            
        except InvalidFileException:
            error_msg = "Invalid Excel file format"
            logger.error(error_msg)
            return False, error_msg
        except Exception as e:
            error_msg = f"Error loading workbook: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def get_sheet_names(self) -> List[str]:
        """
        Get list of available sheet names.
        
        Returns:
            List[str]: List of sheet names
        """
        if self.workbook:
            return self.workbook.sheetnames
        return []
    
    def set_active_sheet(self, sheet_name: str) -> Tuple[bool, str]:
        """
        Set the active sheet and attempt to detect columns.
        
        Args:
            sheet_name (str): Name of the sheet to use
            
        Returns:
            Tuple[bool, str]: (success, error_message)
        """
        if not self.workbook:
            return False, "No workbook loaded"
            
        if sheet_name not in self.workbook.sheetnames:
            return False, f"Sheet '{sheet_name}' not found"
            
        self.sheet = self.workbook[sheet_name]
        
        # Check if sheet is empty
        if self.sheet.max_row < 2:  # Need at least header row and one data row
            return False, "Selected sheet appears to be empty or contains only headers"
            
        # Attempt to detect columns from header row
        self.columns = self._detect_columns()
        if not self.columns:
            return False, "Could not detect required columns (room and SSID) in the header row"
            
        return True, ""
            
    def _detect_columns(self) -> Optional[ExcelColumns]:
        """
        Automatically detect relevant columns from the header row (first row).
        
        Returns:
            Optional[ExcelColumns]: Column indices if found
        """
        if not self.sheet:
            return None
            
        column_indices = {
            'room': None,
            'ssid': None,
            'password': None,
            'encryption': None,
            'property': None
        }
        
        # Get header row (first row)
        header_row = list(self.sheet.iter_rows(min_row=1, max_row=1))[0]
        
        # Search header row for column names
        for idx, cell in enumerate(header_row):
            if not cell.value:
                continue
                
            cell_value = str(cell.value).lower().strip()
            
            # Check for exact matches first
            for col_type, keywords in self.COLUMN_KEYWORDS.items():
                if cell_value in keywords:
                    column_indices[col_type] = idx
                    break
                    
            # If no exact match, check for partial matches
            if all(v is not None for v in column_indices.values()):
                continue
                
            for col_type, keywords in self.COLUMN_KEYWORDS.items():
                if column_indices[col_type] is not None:
                    continue
                if any(keyword in cell_value for keyword in keywords):
                    column_indices[col_type] = idx
                    break
        
        # Log found and missing columns
        found_cols = [k for k, v in column_indices.items() if v is not None]
        missing_cols = [k for k, v in column_indices.items() if v is None]
        logger.info(f"Found columns in header row: {', '.join(found_cols)}")
        if missing_cols:
            logger.warning(f"Missing columns in header row: {', '.join(missing_cols)}")
        
        # Verify required columns are found
        if column_indices['room'] is None or column_indices['ssid'] is None:
            logger.error("Required columns 'room' and 'ssid' not found in header row")
            return None
            
        return ExcelColumns(
            room=column_indices['room'],
            ssid=column_indices['ssid'],
            password=column_indices.get('password'),
            encryption=column_indices.get('encryption'),
            property_type=column_indices.get('property')
        )

    def get_room_data(self, room_number: str) -> Optional[WiFiCredentials]:
        """
        Get WiFi credentials for a specific room.
        
        Args:
            room_number (str): Room number to search for
            
        Returns:
            Optional[WiFiCredentials]: Room's WiFi credentials if found
        """
        if not self.sheet or not self.columns:
            return None
            
        room_number = str(room_number).strip().upper()
        
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if str(row[self.columns.room]).strip().upper() == room_number:
                return WiFiCredentials(
                    ssid=str(row[self.columns.ssid]).strip(),
                    password=str(row[self.columns.password]).strip() if self.columns.password is not None and row[self.columns.password] else None,
                    encryption=str(row[self.columns.encryption]).strip() if self.columns.encryption is not None and row[self.columns.encryption] else "WPA2",
                    property_type=str(row[self.columns.property_type]).strip() if self.columns.property_type is not None and row[self.columns.property_type] else None
                )
        
        logger.warning(f"Room {room_number} not found")
        return None
        
    def get_all_rooms(self) -> List[WiFiCredentials]:
        """
        Get WiFi credentials for all rooms in the sheet.
        
        Returns:
            List[WiFiCredentials]: List of all room credentials
        """
        if not self.sheet or not self.columns:
            return []
            
        credentials = []
        
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if not row[self.columns.room] or not row[self.columns.ssid]:
                continue
                
            cred = WiFiCredentials(
                ssid=str(row[self.columns.ssid]).strip(),
                password=str(row[self.columns.password]).strip() if self.columns.password is not None and row[self.columns.password] else None,
                encryption=str(row[self.columns.encryption]).strip() if self.columns.encryption is not None and row[self.columns.encryption] else "WPA2",
                property_type=str(row[self.columns.property_type]).strip() if self.columns.property_type is not None and row[self.columns.property_type] else None
            )
            credentials.append(cred)
            
        return credentials
        
    def set_columns_manually(self, column_indices: Dict[str, int]) -> bool:
        """
        Manually set column indices.
        
        Args:
            column_indices (Dict[str, int]): Dictionary mapping column names to indices
            
        Returns:
            bool: True if successful
        """
        try:
            self.columns = ExcelColumns(
                room=column_indices['room'],
                ssid=column_indices['ssid'],
                password=column_indices.get('password'),
                encryption=column_indices.get('encryption'),
                property_type=column_indices.get('property_type')
            )
            return True
        except Exception as e:
            logger.error(f"Error setting columns manually: {str(e)}")
            return False